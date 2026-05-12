#!/usr/bin/env python3
"""Parse a Green Tea Sole Company QuickBooks "Transaction Detail by Account" xlsx
and emit gl.gl_entries INSERT statements for the namkhan-pms Supabase project.

Rewritten 2026-05-09 by data-import agent. Replaces broken predecessor.

Schema gotchas (verified against information_schema):
- amount_usd, has_class are GENERATED → MUST NOT appear in INSERT column list.
- upload_id, account_id, class_id, txn_date, period_yyyymm, fiscal_year,
  txn_currency, debit_usd, credit_usd are NOT NULL.
- account_id is FK to gl.accounts; only 6xxxxx and 7xxxxx pass the FK.
- class_id is plain text but constrained to:
  not_specified | undistributed | fb | rooms | spa | transport | imekong | activities | retail
- source_row_hash is text(16) sha256 prefix; idempotency via NOT EXISTS.

Output: a single .sql file with one INSERT … SELECT … WHERE NOT EXISTS … per row.
Caller chooses chunk size when feeding to mcp__supabase__execute_sql.

Usage:
  python3 scripts/load_txn_2026.py --xlsx <path> --upload-id <uuid> --out <sql_path>
                                    [--memo-prefix "[2025-RECON] "]
                                    [--min-date 2026-05-02]   # inclusive
                                    [--property-id 260955]
"""
from __future__ import annotations
import argparse, hashlib, os, re, sys
from datetime import datetime, date
import openpyxl

CLASS_MAP = {
    'F&B': 'fb', 'F & B': 'fb', 'FB': 'fb', 'fb': 'fb',
    'Undistributed': 'undistributed', 'undistributed': 'undistributed',
    'Rooms': 'rooms', 'rooms': 'rooms',
    'Spa': 'spa', 'spa': 'spa',
    'Transport': 'transport', 'transport': 'transport',
    'IMekong': 'imekong', 'iMekong': 'imekong', 'IMEKONG': 'imekong', 'imekong': 'imekong',
    'Activities': 'activities', 'activities': 'activities',
    'Retail': 'retail', 'retail': 'retail',
    'Not specified': 'not_specified', 'Not Specified': 'not_specified',
    'not_specified': 'not_specified',
}
VALID_CLASSES = {'not_specified','undistributed','fb','rooms','spa','transport','imekong','activities','retail'}
SECTION_PAT = re.compile(r'^(\d{6})\s+(.+)$')
FX_LAK_PER_USD = 21800
PROPERTY_ID_DEFAULT = 260955


def sql_str(v, *, null_cast: str = '::text'):
    if v is None:
        return 'NULL' + null_cast
    s = str(v)
    return "'" + s.replace("'", "''") + "'"


def normalise_class(raw) -> str:
    if raw is None:
        return 'not_specified'
    s = str(raw).strip()
    if not s:
        return 'not_specified'
    if s in CLASS_MAP:
        return CLASS_MAP[s]
    sl = s.lower()
    for k, v in CLASS_MAP.items():
        if k.lower() == sl:
            return v
    if sl in VALID_CLASSES:
        return sl
    return 'not_specified'


def parse_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        s = v.strip()
        for fmt in ('%d/%m/%Y', '%m/%d/%Y', '%Y-%m-%d'):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                pass
    return None


def parse(path: str, upload_id: str, property_id: int, memo_prefix: str | None,
          min_date: date | None) -> tuple[list[dict], dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Sheet1']
    section_account = None
    rows = []
    skipped = {
        'no_section': 0,
        'no_date': 0,
        'no_amount': 0,
        'amount_zero': 0,
        'account_not_67': 0,
        'before_min_date': 0,
        'subtotal_or_summary': 0,
    }
    skipped_accts = {}
    classes_seen = {}
    src_file = os.path.basename(path)

    # row 1 = company name, 2 = title, 3 = period, 4 = blank, 5 = headers, 6+ = data
    for r in range(6, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value

        # Section header rows: col 1 has "<acct_id> <name>", col 2 empty
        if a is not None:
            a_s = str(a).strip()
            if a_s:
                m = SECTION_PAT.match(a_s)
                if m:
                    section_account = m.group(1)
                    continue
                # Not a section, not a transaction (cols 2..N empty) → likely "Total ...", "Beginning Balance", or section header like "Income"
                if b is None:
                    skipped['subtotal_or_summary'] += 1
                    continue

        # Transaction row: col 1 should be empty, col 2 = date
        if b is None:
            continue
        txn_date = parse_date(b)
        if txn_date is None:
            skipped['no_date'] += 1
            continue
        if min_date is not None and txn_date < min_date:
            skipped['before_min_date'] += 1
            continue
        if section_account is None:
            skipped['no_section'] += 1
            continue

        # Filter: only 6xxxxx and 7xxxxx accounts
        if not (section_account.startswith('6') or section_account.startswith('7')):
            skipped['account_not_67'] += 1
            skipped_accts[section_account] = skipped_accts.get(section_account, 0) + 1
            continue

        txn_type = ws.cell(r, 3).value
        txn_number = ws.cell(r, 4).value
        party = ws.cell(r, 5).value
        location = ws.cell(r, 6).value
        klass = ws.cell(r, 7).value
        desc = ws.cell(r, 8).value
        split = ws.cell(r, 9).value
        amount = ws.cell(r, 10).value

        if amount is None:
            skipped['no_amount'] += 1
            continue
        try:
            amt = float(amount)
        except (TypeError, ValueError):
            skipped['no_amount'] += 1
            continue
        if amt == 0:
            skipped['amount_zero'] += 1
            continue

        class_id = normalise_class(klass)
        classes_seen[class_id] = classes_seen.get(class_id, 0) + 1

        # Sign convention: positive = debit, negative = credit
        if amt > 0:
            debit_usd = round(amt, 4)
            credit_usd = 0.0
        else:
            debit_usd = 0.0
            credit_usd = round(-amt, 4)
        amount_lak = round(amt * FX_LAK_PER_USD, 4)

        period_yyyymm = txn_date.strftime('%Y-%m')
        fiscal_year = txn_date.year

        memo_parts = []
        if desc:
            memo_parts.append(str(desc).strip())
        if location:
            memo_parts.append(f"loc={str(location).strip()}")
        if split:
            memo_parts.append(f"split={str(split).strip()}")
        memo = ' | '.join(memo_parts) if memo_parts else None
        if memo_prefix:
            memo = (memo_prefix + (memo or ''))

        # Hash includes everything that distinguishes the row
        hsrc = '|'.join([
            src_file,
            str(r),
            section_account,
            txn_date.isoformat(),
            str(txn_type or ''),
            str(txn_number or ''),
            str(party or ''),
            str(klass or ''),
            str(desc or ''),
            str(split or ''),
            f"{amt:.4f}",
        ])
        h = hashlib.sha256(hsrc.encode('utf-8')).hexdigest()[:16]

        rows.append({
            'upload_id': upload_id,
            'qb_txn_id': None,
            'qb_txn_type': str(txn_type)[:50] if txn_type else None,
            'qb_txn_number': str(txn_number) if txn_number not in (None, '') else None,
            'txn_date': txn_date.isoformat(),
            'period_yyyymm': period_yyyymm,
            'fiscal_year': fiscal_year,
            'account_id': section_account,
            'class_id': class_id,
            'customer_name': str(party).strip() if party not in (None, '') else None,
            'memo': memo,
            'debit_usd': debit_usd,
            'credit_usd': credit_usd,
            'txn_currency': 'USD',
            'txn_amount_native': round(amt, 4),
            'fx_rate_used': FX_LAK_PER_USD,
            'amount_lak': amount_lak,
            'source_row_index': r,
            'source_row_hash': h,
            'property_id': property_id,
        })

    stats = {
        'rows': len(rows),
        'skipped': skipped,
        'skipped_accounts': dict(sorted(skipped_accts.items(), key=lambda x: -x[1])[:20]),
        'classes_seen': classes_seen,
    }
    if rows:
        stats['date_min'] = min(r['txn_date'] for r in rows)
        stats['date_max'] = max(r['txn_date'] for r in rows)
        stats['accounts_distinct'] = len(set(r['account_id'] for r in rows))
    return rows, stats


def emit_sql(rows: list[dict], out_path: str, batch_size: int = 200):
    """Emit one INSERT … SELECT … FROM (VALUES (…),(…),…) v(…) WHERE NOT EXISTS … per batch."""
    cols = ['upload_id','qb_txn_id','qb_txn_type','qb_txn_number','txn_date',
            'period_yyyymm','fiscal_year','account_id','class_id','customer_name','memo',
            'debit_usd','credit_usd','txn_currency','txn_amount_native','fx_rate_used',
            'amount_lak','source_row_index','source_row_hash','property_id']
    # entry_id is added as gen_random_uuid() in the INSERT column list
    with open(out_path, 'w') as f:
        for start in range(0, len(rows), batch_size):
            batch = rows[start:start + batch_size]
            value_tuples = []
            for row in batch:
                vals = [
                    f"'{row['upload_id']}'::uuid",
                    sql_str(row['qb_txn_id']),
                    sql_str(row['qb_txn_type']),
                    sql_str(row['qb_txn_number']),
                    f"'{row['txn_date']}'::date",
                    sql_str(row['period_yyyymm']),
                    f"{row['fiscal_year']}::int",
                    sql_str(row['account_id']),
                    sql_str(row['class_id']),
                    sql_str(row['customer_name']),
                    sql_str(row['memo']),
                    f"{row['debit_usd']}::numeric",
                    f"{row['credit_usd']}::numeric",
                    sql_str(row['txn_currency']),
                    f"{row['txn_amount_native']}::numeric",
                    f"{row['fx_rate_used']}::numeric",
                    f"{row['amount_lak']}::numeric",
                    f"{row['source_row_index']}::int",
                    sql_str(row['source_row_hash']),
                    f"{row['property_id']}::bigint",
                ]
                value_tuples.append('(' + ', '.join(vals) + ')')
            f.write(
                "INSERT INTO gl.gl_entries (entry_id, " + ', '.join(cols) + ") "
                "SELECT gen_random_uuid(), " + ', '.join('v.' + c for c in cols) + " "
                "FROM (VALUES " + ', '.join(value_tuples) + ") AS v(" + ', '.join(cols) + ") "
                "WHERE NOT EXISTS (SELECT 1 FROM gl.gl_entries g "
                "WHERE g.source_row_hash = v.source_row_hash AND g.property_id = v.property_id);\n"
            )


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--xlsx', required=True)
    ap.add_argument('--upload-id', required=True)
    ap.add_argument('--out', required=True)
    ap.add_argument('--memo-prefix', default=None)
    ap.add_argument('--min-date', default=None, help='YYYY-MM-DD inclusive')
    ap.add_argument('--property-id', type=int, default=PROPERTY_ID_DEFAULT)
    args = ap.parse_args()

    min_d = datetime.strptime(args.min_date, '%Y-%m-%d').date() if args.min_date else None
    rows, stats = parse(args.xlsx, args.upload_id, args.property_id, args.memo_prefix, min_d)
    emit_sql(rows, args.out)
    import json
    print(json.dumps(stats, indent=2, default=str))
    print(f"WROTE {len(rows)} INSERTs -> {args.out}")


if __name__ == '__main__':
    main()
