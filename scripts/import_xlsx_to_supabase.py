#!/usr/bin/env python3
"""Parse Green Tea P&L and transactions xlsx files into JSON for SQL insertion.

Used by Claude data-import agent on 2026-05-09 to import five xlsx files into
Supabase project kpenyneooigsyuuomgct (namkhan-pms).
"""
import openpyxl, hashlib, json, re, os, sys
from datetime import datetime

DESKTOP = "/Users/paulbauer/Desktop"
OUT_DIR = "/tmp"

def fhash(p):
    h = hashlib.sha256()
    with open(p, 'rb') as f:
        for chunk in iter(lambda: f.read(8192), b''):
            h.update(chunk)
    return h.hexdigest()[:16]

def parse_pl_with_months(path, year):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"]
    header_row = None
    for r in range(1, 10):
        v = ws.cell(r, 2).value
        if isinstance(v, str) and ('Jan' in v or 'Apr' in v) and str(year) in v:
            header_row = r
            break
    if not header_row:
        return None
    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column+1)]
    month_cols = {}
    month_to_num = {'Jan':1,'Feb':2,'Mar':3,'Apr':4,'May':5,'Jun':6,
                    'Jul':7,'Aug':8,'Sep':9,'Oct':10,'Nov':11,'Dec':12}
    month_pat = re.compile(r'^([A-Z][a-z]{2})\s+(\d{4})')
    for i, h in enumerate(headers, start=1):
        if not isinstance(h, str): continue
        m = month_pat.match(h.strip())
        if m and m.group(1) in month_to_num:
            yyyymm = f"{m.group(2)}-{month_to_num[m.group(1)]:02d}"
            month_cols[i] = yyyymm
    rows = []
    section = None
    for r in range(header_row+1, ws.max_row+1):
        a = ws.cell(r, 1).value
        if not a: continue
        a_str = str(a).strip()
        if not a_str: continue
        if not re.match(r'^\d{6}', a_str):
            if a_str.startswith('Total'):
                continue
            if a_str in ('Income','Cost of Goods Sold','Cost of Sales','Expenses',
                        'Other Income','Other Expenses','Gross Profit',
                        'Net Operating Income','Net Other Income','Net Income'):
                section = a_str
            continue
        m = re.match(r'^(\d{6})\s+(.+)$', a_str)
        if not m: continue
        acct_id = m.group(1)
        for col, yyyymm in month_cols.items():
            v = ws.cell(r, col).value
            if v is None: continue
            try:
                amount = float(v)
            except (TypeError, ValueError):
                continue
            if amount == 0: continue
            rows.append({
                'period_yyyymm': yyyymm,
                'account_id': acct_id,
                'amount_usd': round(amount, 4),
                'section': section,
            })
    return rows

def parse_pl_class(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"]
    classes_row = 5
    headers = [ws.cell(classes_row, c).value for c in range(1, ws.max_column+1)]
    class_map = {
        'Activities': 'activities', 'F&B': 'fb', 'IMekong': 'imekong',
        'Retail': 'retail', 'Rooms': 'rooms', 'Spa': 'spa',
        'Transport': 'transport', 'Undistributed': 'undistributed',
        'Not specified': 'not_specified',
    }
    class_cols = {}
    for i, h in enumerate(headers, start=1):
        if isinstance(h, str) and h.strip() in class_map:
            class_cols[i] = class_map[h.strip()]
    rows = []
    section = None
    for r in range(6, ws.max_row+1):
        a = ws.cell(r, 1).value
        if not a: continue
        a_str = str(a).strip()
        if not a_str: continue
        if not re.match(r'^\d{6}', a_str):
            if a_str.startswith('Total'): continue
            if a_str in ('Income','Cost of Goods Sold','Cost of Sales','Expenses',
                        'Other Income','Other Expenses','Gross Profit',
                        'Net Operating Income','Net Other Income','Net Income'):
                section = a_str
            continue
        m = re.match(r'^(\d{6})\s+(.+)$', a_str)
        if not m: continue
        acct_id = m.group(1)
        for col, class_id in class_cols.items():
            v = ws.cell(r, col).value
            if v is None: continue
            try: amount = float(v)
            except (TypeError, ValueError): continue
            if amount == 0: continue
            rows.append({
                'account_id': acct_id,
                'class_id': class_id,
                'amount_usd': round(amount, 4),
                'section': section,
            })
    return rows

def parse_transactions(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"]
    rows = []
    current_section = None
    section_pat = re.compile(r'^(\d{6})\s+(.+)$')
    src_file = os.path.basename(path)
    for r in range(6, ws.max_row+1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        if a and isinstance(a, str) and section_pat.match(a.strip()):
            m = section_pat.match(a.strip())
            current_section = m.group(1)
            continue
        if a and isinstance(a, str) and (a.strip().startswith('Total') or a.strip().startswith('Beginning Balance')):
            continue
        if not b: continue
        txn_date = None
        if isinstance(b, datetime):
            txn_date = b.date().isoformat()
        elif isinstance(b, str):
            for fmt in ('%d/%m/%Y','%m/%d/%Y','%Y-%m-%d'):
                try:
                    txn_date = datetime.strptime(b.strip(), fmt).date().isoformat()
                    break
                except ValueError:
                    pass
        if not txn_date: continue
        txn_type = ws.cell(r, 3).value or ''
        txn_number = ws.cell(r, 4).value
        party = ws.cell(r, 5).value
        location = ws.cell(r, 6).value
        klass = ws.cell(r, 7).value
        desc = ws.cell(r, 8).value
        split = ws.cell(r, 9).value
        amount = None
        for col in (10, 11):
            v = ws.cell(r, col).value
            if v is not None:
                try:
                    amount = float(v)
                    break
                except (TypeError, ValueError):
                    pass
        if amount is None: continue
        line_account = None
        if split and isinstance(split, str):
            mm = re.match(r'^(\d{6})\s+', split.strip())
            if mm: line_account = mm.group(1)
        rows.append({
            'txn_date': txn_date,
            'txn_type': str(txn_type)[:50] if txn_type else 'Unknown',
            'txn_number': str(txn_number) if txn_number else None,
            'section_account': current_section,
            'line_account': line_account,
            'party_name': str(party) if party else None,
            'location': str(location) if location else None,
            'class': str(klass) if klass else None,
            'description': str(desc) if desc else None,
            'amount_native': round(amount, 2),
            'source_row': r,
            'source_file': src_file,
        })
    return rows

def main():
    target = sys.argv[1] if len(sys.argv) > 1 else 'all'

    jobs = []
    if target in ('all','pl_2026'):
        jobs.append(('pl_2026',
                     "Green Tea Sole Company Limited_Profit and Loss - 2026.xlsx",
                     lambda f: parse_pl_with_months(f, 2026)))
    if target in ('all','pl_2025'):
        jobs.append(('pl_2025',
                     "Green Tea Sole Company Limited_Profit and Loss - 2025 .xlsx",
                     lambda f: parse_pl_with_months(f, 2025)))
    if target in ('all','pl_class_2026'):
        jobs.append(('pl_class_2026',
                     "Green Tea Sole Company Limited_P_L CLass 26.xlsx",
                     parse_pl_class))
    if target in ('all','txn_2025'):
        jobs.append(('txn_2025',
                     "Green Tea Sole Company Limited_Transaction Detail by Account 2025.xlsx",
                     parse_transactions))
    if target in ('all','txn_2026'):
        jobs.append(('txn_2026',
                     "Green Tea Sole Company Limited_Transaction Detail by Account  2026.xlsx",
                     parse_transactions))

    for key, fname, parser in jobs:
        f = os.path.join(DESKTOP, fname)
        if not os.path.exists(f):
            print(f"SKIP {key}: file not found {f}")
            continue
        rows = parser(f)
        h = fhash(f)
        out = {'rows': rows, 'hash': h, 'file': fname}
        out_path = os.path.join(OUT_DIR, f"{key}.json")
        with open(out_path, 'w') as o:
            json.dump(out, o)
        print(f"{key}: {len(rows)} rows | hash={h} | -> {out_path}")
        if rows and 'period_yyyymm' in rows[0]:
            print(f"  periods: {sorted(set(r['period_yyyymm'] for r in rows))}")
        if rows and 'txn_date' in rows[0]:
            print(f"  date range: {min(r['txn_date'] for r in rows)} .. {max(r['txn_date'] for r in rows)}")
        if rows and 'class_id' in rows[0]:
            print(f"  classes: {sorted(set(r['class_id'] for r in rows))}")
            print(f"  income total: {sum(r['amount_usd'] for r in rows if r['section']=='Income'):.2f}")

if __name__ == '__main__':
    main()
